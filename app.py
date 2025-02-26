import pyautogui
import time
import webbrowser
from urllib.parse import quote
import openpyxl
import tkinter as tk
from tkinter import ttk
import threading
from datetime import datetime

# Variável global para controlar o envio
envio_ativo = threading.Event()

def enviar_mensagem_com_enter(cliente, mensagem):
    try:
        telefone_formatado = cliente["telefone"]
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone_formatado}&text={quote(mensagem)}'
        webbrowser.open(link_mensagem_whatsapp)
        time.sleep(20)
        pyautogui.press('enter')
        time.sleep(5)
    except Exception as e:
        print(f"Erro ao enviar para {cliente['nome']}: {e}")

def criar_mensagem(cliente):
    grupo = cliente["grupo"].lower()
    nome = cliente["nome"]
    empresa = cliente["empresa"]
    inicio = cliente["inicio"]
    mensagens = {
        "clientes": f"{inicio}, {nome}, da empresa {empresa}, Como está o uso do sistema? Tudo certo? Algo a relatar? Entre em contato para mais informações no (54) 9 9104 1029. Prestamos suporte Técnico rápido e ativo, mas Caso você não queira mais receber informações sobre nossos serviços e produtos, basta nos enviar a palavra: *SAIR*.",
        #
        "prospects": f"{inicio}, {nome}, da empresa {empresa}, Nós da *GDI Informatica*, trabalhamos com um *Sistema de Gestão: Simples e Prático*, e montamos a cobrança encima das ferramentas que realmente for utilizar. * Mas Caso você não queira mais receber informações, sobre nossos serviços e produtos, basta nos enviar a palavra: SAIR*.",
        #
        "pos": f"{inicio}, {nome}, da empresa {empresa}, Como está o uso do Aplicativo na Maquininha de Cartões, tudo certo, Algo a Relatar?. Entre em contato para mais informações no (54) 9 9104 1029. Prestamos Suporte Técnico Rápido e Ativo, mas Caso você não queira mais receber informações sobre nossos serviços, basta nos enviar a palavra: *SAIR*.",
        #
        "em negociacao": f"{inicio}, {nome}, da empresa {empresa}, Nós da *GDI Informatica*, trabalhamos com um *Sistema de Gestão: Simples e Prático*, ja tinhamos conversado algo, agora temos *Novidades* e *Promoções* exclusivas para você nesse novo Ano. Entre em contato comigo, o Glaucio! *Mas Caso você não queira mais receber informações, sobre nossos serviços, basta nos enviar a palavra: SAIR*.",
        #
        "parceiros": f"{inicio}, {nome}, da empresa {empresa}, Nós da *GDI Informatica*, trabalhamos com um *Sistema de Gestão: Simples e Prático*, para reafirmar nossa parceria, agora temos *Novidades* e *Promoções* exclusivas para seus Clientes nesse Novo Ano. Entre em contato comigo, o Glaucio! *Mas Caso você não queira mais receber informações, sobre nossos serviços, basta nos enviar a palavra: SAIR*.",
        #
        "contadores": f"{inicio}, {nome}, da empresa {empresa}, Visitei o seu Escritório pessoalmente, e estamos entrando em contato novamente, para dar uma solução pratica e simples para seu cliente, referente a sistema de Gestão. Entre em contato!"    }
        #
    return mensagens.get(grupo, f"Olá {nome}, da empresa {empresa}, Indique-nos para seus Clientes!")

def ler_dados_planilha(pagina_selecionada, grupo_selecionado):
    try:
        workbook = openpyxl.load_workbook('Mala_Whats.xlsx')
        aba = workbook[pagina_selecionada]
        clientes = []
        for linha in aba.iter_rows(min_row=2, values_only=True):
            if linha[2]:
                cliente = {
                    "empresa": linha[0],
                    "nome": linha[1],
                    "telefone": linha[2],
                    "inicio": linha[3] if linha[3] else "Indefinido",
                    "grupo": linha[4].strip().lower() if linha[4] else "indefinido"
                }
                if grupo_selecionado == "Todos" or cliente["grupo"] == grupo_selecionado.lower():
                    clientes.append(cliente)
        return clientes
    except Exception as e:
        print(f"Erro ao ler a planilha: {e}")
        return []

def iniciar_envio(dados):
    envio_ativo.set()
    for cliente, mensagem in dados:
        if not envio_ativo.is_set():
            break
        enviar_mensagem_com_enter(cliente, mensagem)
        time.sleep(120)

def setup_gui():
    janela = tk.Tk()
    janela.title('Envio Automático de Mensagens WhatsApp')
    janela.geometry("1550x600")
    
    abas = openpyxl.load_workbook('Mala_Whats.xlsx').sheetnames
    pagina_var = tk.StringVar(value=abas[0])
    grupo_var = tk.StringVar(value="Todos")
    
    tk.Label(janela, text="Selecione a Página do Excel:").pack()
    tk.OptionMenu(janela, pagina_var, *abas).pack()
    tk.Label(janela, text="Selecione o TIPO DE MENSAGEM:").pack()
    tk.OptionMenu(janela, grupo_var, "Todos", "Clientes", "Prospects", "Parceiros", "Em Negociacao", "Contadores", "Pos").pack()
    
    frame = ttk.Frame(janela)
    frame.pack(fill=tk.BOTH, expand=True)
    
    tree = ttk.Treeview(frame, columns=("Nome", "Telefone", "Mensagem"), show="headings")

# Definição dos cabeçalhos
    tree.heading("Nome", text="Nome")
    tree.heading("Telefone", text="Telefone")
    tree.heading("Mensagem", text="Mensagem")

# Ajuste do tamanho das colunas
    tree.column("Nome", width=10, anchor="w")  # Nome menor
    tree.column("Telefone", width=10, anchor="center")  # Telefone menor
    tree.column("Mensagem", width=500, anchor="w")  # Mensagem maior

    tree.pack(fill=tk.BOTH, expand=True)

    
    def carregar_dados():
        for row in tree.get_children():
            tree.delete(row)
        clientes = ler_dados_planilha(pagina_var.get(), grupo_var.get())
        for cliente in clientes:
            mensagem = criar_mensagem(cliente)
            tree.insert("", tk.END, values=(cliente['nome'], cliente['telefone'], mensagem))
    
    tk.Button(janela, text="Carregar Dados", command=carregar_dados).pack()
    
    def editar_mensagem():
        selected_item = tree.selection()
        if selected_item:
            item = tree.item(selected_item)
            top = tk.Toplevel(janela)
            top.title("Editar Mensagem")
            tk.Label(top, text="Mensagem:").pack()
            texto = tk.Text(top, height=5, width=50)
            texto.insert(tk.END, item['values'][2])
            texto.pack()
            def salvar():
                tree.item(selected_item, values=(item['values'][0], item['values'][1], texto.get("1.0", tk.END).strip()))
                top.destroy()
            tk.Button(top, text="Salvar", command=salvar).pack()
    
    tk.Button(janela, text="Editar Mensagem", command=editar_mensagem).pack()
    
    def iniciar_envio_thread():
        dados_envio = []
        
        for item in tree.get_children():
            valores = tree.item(item)["values"]
            cliente = {
                "nome": valores[0],
                "telefone": valores[1]
            }
            mensagem = valores[2]
            dados_envio.append((cliente, mensagem))
        
        if not dados_envio:
            print("Nenhum dado carregado para envio.")
            return

        print("Iniciando envio de mensagens...")
        thread_envio = threading.Thread(target=iniciar_envio, args=(dados_envio,), daemon=True)
        thread_envio.start()
    
    tk.Button(janela, text="Iniciar Envio", command=iniciar_envio_thread, bg="orange", fg="white").pack()
    tk.Button(janela, text="STOP Envio", command=lambda: envio_ativo.clear()).pack()
    tk.Button(janela, text="Fechar", command=janela.destroy).pack()
    
    janela.mainloop()

setup_gui()

