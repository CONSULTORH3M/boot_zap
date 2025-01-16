import pyautogui
import time
import webbrowser
from urllib.parse import quote
import openpyxl
import tkinter as tk
import threading
from datetime import datetime

# Variável global para controlar o envio
envio_ativo = threading.Event()

# Função para enviar a mensagem
def enviar_mensagem_com_enter(cliente, mensagem, listbox):
    try:
        telefone_formatado = cliente["telefone"]
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone_formatado}&text={quote(mensagem)}'
        webbrowser.open(link_mensagem_whatsapp)

        # Aumentando o tempo para o WhatsApp carregar
        time.sleep(20)  # Aumenta o tempo de espera para o WhatsApp carregar corretamente

        # Aguardar mais um pouco para garantir que a página está completamente carregada
        time.sleep(10)  # Espera mais 10 segundos para garantir que a interface de envio esteja pronta

        # Envia a mensagem
        pyautogui.press('enter')  # Envia a mensagem
        time.sleep(10)  # Espera 10 segundos para garantir que a mensagem foi enviada

        # Fecha a aba do navegador
        #pyautogui.hotkey('ctrl', 'w')  # Fecha a aba do navegador

        # Registra o horário do envio
        hora_envio = datetime.now().strftime("%H:%M:%S")
        listbox.insert(tk.END, f"[{hora_envio}] Mensagem enviada para {cliente['nome']} ({telefone_formatado}).\n")
        listbox.yview(tk.END)
    except Exception as e:
        listbox.insert(tk.END, f"Erro ao enviar para {cliente['nome']}: {e}\n")
        listbox.yview(tk.END)

# Função para criar a mensagem baseada no grupo
def criar_mensagem(cliente):
    grupo = cliente["grupo"].lower()  # Converte o grupo para minúsculo para garantir a comparação correta
    nome = cliente["nome"]
    empresa = cliente["empresa"]
    inicio = cliente["inicio"]

    # Mensagem para o grupo "ativos"
    if grupo == "ativos":
        return f"{inicio}, {nome}, da empresa {empresa}, Como está o uso do sistema? Tudo certo? Algo a relatar? Entre em contato para mais informações no (54) 9 9104 1029. Prestamos suporte Técnico rápido e ativo, mas Caso você não queira mais receber informações sobre nossos serviços e produtos, basta nos enviar a palavra: *SAIR*."
    
    # Mensagem para o grupo "prospects"
    elif grupo == "prospects":
        return f"{inicio}, {nome}, da empresa {empresa}, Nós da *GDI Informatica*, trabalhamos com um *Sistema de Gestão: Simples e Prático*, e montamos a cobrança encima das ferramentas que realmente for utilizar. * Mas Caso você não queira mais receber informações, sobre nossos serviços e produtos, basta nos enviar a palavra: SAIR*."

    # Mensagem para o grupo "POS"
    elif grupo == "pos":
        return f"{inicio}, {nome}, da empresa {empresa}, Como está o usa o Aplicativo, tudo Certo, Algo a relatar. Entre em contato para mais informações no (54) 9 9104 1029. Prestamos suporte Técnico rápido e ativo, mas Caso você não queira mais receber informações sobre nossos serviços, basta nos enviar a palavra: *SAIR*."

    # Mensagem para o grupo "clientes"
    elif grupo == "em negociacao":
        return f"{inicio}, {nome}, da empresa {empresa}, Nós da *GDI Informatica*, trabalhamos com um *Sistema de Gestão: Simples e Prático*, 
        ja tinhamos conversado algo, agora temos novidades e promoções exclusivas para você nesse novo Ano. Entre em contato comigo, o Glaucio!   *Mas Caso você não queira mais receber informações, sobre nossos serviços, basta nos enviar a palavra: SAIR*."    

    # Mensagem para o grupo "clientes"
    elif grupo == "contadores":
        return f"{inicio}, {nome}, da empresa {empresa}, Visitei o seu Escritório pessoalmente, e estamos entrando em contato novamente, para dar uma solução pratica e simples para seu cliente, referente a sistema de Gestão. Entre em contato!"
    
    # Caso o grupo não corresponda a nenhum dos definidos, a mensagem será padrão
    else:
        return f"Olá {nome}, da empresa {empresa}, *Indique-nos* para seus Clientes. Garantimos que eles serão muito bem antendidos. Entre em contato comigo, o Glaucio e comprove!!"
import re

import re

def normalizar_telefone(telefone):
    # Remove caracteres não numéricos
    telefone = re.sub(r'\D', '', telefone)
    
    # Verifica se o número começa com o código do país (55)
    if telefone.startswith("55"):
        telefone = telefone[2:]  # Remove o código do país para simplificar
    
    # Verifica se o número tem DDD e número de celular com 8 ou 9 dígitos
    if len(telefone) == 10:  # Formato DDD + 8 dígitos
        telefone = telefone[:2] + "9" + telefone[2:]  # Adiciona o dígito 9
    elif len(telefone) == 11:  # Formato DDD + 9 dígitos (já correto)
        pass
    else:
        return None  # Número inválido
    
    # Adiciona o código do país (55) no início
    telefone = "55" + telefone
    return telefone

# Função para ler os dados da planilha
def ler_dados_planilha(pagina_selecionada, grupo_selecionado):
    try:
        workbook = openpyxl.load_workbook('mala_whats.xlsx')
        aba = workbook[pagina_selecionada]
        clientes = []

        for linha in aba.iter_rows(min_row=2, values_only=True):
            if linha[2]:  # Verifica se o telefone está preenchido
                telefone_normalizado = normalizar_telefone(str(linha[2]))
                if telefone_normalizado:  # Verifica se o telefone é válido
                    cliente = {
                        "empresa": linha[0],
                        "nome": linha[1],
                        "telefone": telefone_normalizado,
                        "inicio": linha[3] if linha[3] else "Indefinido",
                        "grupo": linha[4] if linha[4] else "Indefinido"
                    }
                    if grupo_selecionado == "Todos" or cliente["grupo"].lower() == grupo_selecionado.lower():
                        clientes.append(cliente)
        return clientes
    except Exception as e:
        print(f"Erro ao ler a planilha: {e}")
        return []

# Função para iniciar o envio
def iniciar_envio(listbox, grupo_selecionado, pagina_selecionada):
    envio_ativo.set()  # Ativar o envio
    clientes = ler_dados_planilha(pagina_selecionada, grupo_selecionado)
    if not clientes:
        listbox.insert(tk.END, "Nenhum cliente encontrado para o grupo selecionado.\n")
        listbox.yview(tk.END)
        return

    for cliente in clientes:
        if not envio_ativo.is_set():
            listbox.insert(tk.END, "Envio interrompido pelo usuário.\n")
            listbox.yview(tk.END)
            break

        mensagem = criar_mensagem(cliente)
        listbox.insert(tk.END, f"Enviando a {cliente['nome']} ({cliente['telefone']})...Aguardando 30s para Disparar \n")
        listbox.yview(tk.END)
        enviar_mensagem_com_enter(cliente, mensagem, listbox)
        
        # Intervalo de 2 minutos entre os envios
        for i in range(120):  # Adiciona uma contagem visual do tempo
            if not envio_ativo.is_set():
                break
            time.sleep(1)  # Espera 1 segundo (total: 120 segundos)

# Função para parar o envio
def parar_envio(listbox):
    envio_ativo.clear()
    listbox.insert(tk.END, "Envio interrompido.\n")
    listbox.yview(tk.END)

# Função para obter as abas do Excel
def obter_abas_excel():
    try:
        workbook = openpyxl.load_workbook('mala_whats.xlsx')
        return workbook.sheetnames
    except Exception as e:
        print(f"Erro ao obter abas do Excel: {e}")
        return []

# Função para configurar a interface gráfica
def setup_gui():
    def iniciar_envio_thread():
        grupo_selecionado = grupo_var.get()
        pagina_selecionada = pagina_var.get()
        threading.Thread(target=iniciar_envio, args=(listbox, grupo_selecionado, pagina_selecionada)).start()

    abas = obter_abas_excel()
    if not abas:
        print("Erro: Não foi possível carregar as abas do Excel.")
        return

    janela = tk.Tk()
    janela.title('Envio Automático de Mensagens WhatsApp')
    janela.geometry("600x580")

    # Dropdown para selecionar a aba do Excel
    pagina_var = tk.StringVar(value=abas[0])
    tk.Label(janela, text="Selecione a Página do Excel:").pack(pady=5)
    pagina_dropdown = tk.OptionMenu(janela, pagina_var, *abas)
    pagina_dropdown.pack(pady=5)

    # Dropdown para selecionar o grupo
    grupo_var = tk.StringVar(value="Todos")
    tk.Label(janela, text="Selecione o Modelo Mensagem(Grupo):").pack(pady=5)
    grupo_dropdown = tk.OptionMenu(janela, grupo_var, "Todos", "Ativos", "Prospects", "Parceiros", "Em Negociacao", "Contadores")
    grupo_dropdown.pack(pady=5)

    # Botão para iniciar o envio
    tk.Button(janela, text="Iniciar Envio", command=iniciar_envio_thread, bg="orange", fg="white").pack(pady=10)

    # Área para exibir logs
    listbox = tk.Listbox(janela, width=70, height=15)
    listbox.pack(pady=10)

    # Botão para parar o envio
    tk.Button(janela, text="Parar Envio", command=lambda: parar_envio(listbox)).pack(pady=5)

    # Botão para fechar a aplicação
    tk.Button(janela, text="Fechar", command=lambda: [parar_envio(listbox), janela.destroy()]).pack(pady=5)

    janela.mainloop()

# Executa a interface gráfica
setup_gui()
